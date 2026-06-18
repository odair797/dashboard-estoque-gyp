# -*- coding: utf-8 -*-
import sys
try:
    sys.stdout.reconfigure(encoding='utf-8'); sys.stderr.reconfigure(encoding='utf-8')
except Exception: pass
import os, re
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
WORKSPACE=os.environ.get('GYP_OUT','.'); OUT_XLSX=os.path.join(WORKSPACE,'Analise_Estoque_GYP.xlsx')
TODAY=datetime.today(); DATA_STR=TODAY.strftime('%d/%m/%Y')
PROC_RE=re.compile(r'DOCA|PACKING',re.IGNORECASE)
def em_processo(setor): return bool(PROC_RE.search(str(setor or '')))
DARK_BG='1B2330'; GOLD='C9A961'; WHITE='FFFFFF'
RED_CRIT='C0392B'; ORG_ATEN='D68910'; YEL_PROX='F1C40F'
GREEN_OK='27AE60'; GRAY_BAND='EAECEE'; LINHA_BG='34495E'; NO_DATA='7F8C8D'; PROC_BG='5D6D7E'
thin=Border(left=Side(style='thin',color='B0B0B0'),right=Side(style='thin',color='B0B0B0'),top=Side(style='thin',color='B0B0B0'),bottom=Side(style='thin',color='B0B0B0'))
def title_row(ws,row,ncols,text):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=ncols)
    c=ws.cell(row=row,column=1,value=text); c.font=Font(name='Calibri',size=14,bold=True,color=GOLD)
    c.fill=PatternFill('solid',fgColor=DARK_BG); c.alignment=Alignment(horizontal='center',vertical='center'); ws.row_dimensions[row].height=28
def header_row(ws,row,headers):
    for i,h in enumerate(headers,1):
        c=ws.cell(row=row,column=i,value=h); c.font=Font(name='Calibri',size=10,bold=True,color=WHITE)
        c.fill=PatternFill('solid',fgColor=DARK_BG); c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=thin
    ws.row_dimensions[row].height=36
CRIT_FILL={'VENCIDO':(RED_CRIT,WHITE),'CRÍTICO ≤60d':(RED_CRIT,WHITE),'ATENÇÃO ≤120d':(ORG_ATEN,WHITE),'PRÓXIMO ≤180d':(YEL_PROX,'000000'),'OK':(GREEN_OK,WHITE),'SEM DATA':(NO_DATA,WHITE),'VENCIDO ':(RED_CRIT,WHITE)}
def fmt_data(v):
    if pd.isna(v) or v=='': return ''
    try: return pd.to_datetime(v).strftime('%d/%m/%Y')
    except: return v
def write_row(ws,row,headers,valores):
    band=GRAY_BAND if (row%2==0) else WHITE; idx={h:i for i,h in enumerate(headers,1)}
    for i,v in enumerate(valores,1):
        c=ws.cell(row=row,column=i,value=v); c.font=Font(name='Calibri',size=9); c.border=thin
        c.fill=PatternFill('solid',fgColor=band); c.alignment=Alignment(horizontal='center',vertical='center')
    if 'Descrição' in idx: ws.cell(row=row,column=idx['Descrição']).alignment=Alignment(horizontal='left',vertical='center')
    if 'Quantidade' in idx: ws.cell(row=row,column=idx['Quantidade']).number_format='#,##0.000'
    if 'Dias a Vencer' in idx: ws.cell(row=row,column=idx['Dias a Vencer']).number_format='0'
    if 'Meses a Vencer' in idx: ws.cell(row=row,column=idx['Meses a Vencer']).number_format='0.0'
    if 'Criticidade' in idx:
        cc=idx['Criticidade']; cv=valores[cc-1]
        if cv in CRIT_FILL:
            bg,fg=CRIT_FILL[cv]; c=ws.cell(row=row,column=cc); c.fill=PatternFill('solid',fgColor=bg); c.font=Font(name='Calibri',size=9,bold=True,color=fg)
HDRS_FAM=['SKU','Descrição','Quantidade','Lote Indústria','Data de Fabricação','Local','Setor','Data de Vencimento','Dias a Vencer','Meses a Vencer','Criticidade']
HDRS_PA=['Linha']+HDRS_FAM; HDRS_PROC=['Família']+HDRS_FAM
WID_FAM=[16,42,13,18,16,12,26,14,12,12,16]; WID_PA=[18]+WID_FAM; WID_PROC=[20]+WID_FAM
def rec_padrao(r):
    return [r.get('Código do Produto',''),r.get('Produto',''),float(r.get('Estoque (UN)',0) or 0),r.get('Lote Indústria',''),fmt_data(r.get('Data de Fabricação','')),r.get('Local',''),r.get('Setor',''),fmt_data(r.get('Data de Vencimento','')),(int(r['Dias p/ Vencer']) if pd.notna(r.get('Dias p/ Vencer')) else ''),(float(r['Meses p/ Vencer']) if pd.notna(r.get('Meses p/ Vencer')) else ''),r.get('Criticidade','')]
wb=Workbook(); wb.remove(wb.active); TMP=os.environ.get('GYP_TMP','/tmp'); processo_rows=[]
pa=pd.read_csv(os.path.join(TMP,'fam_PA.csv'))
pa_proc_mask=pa['Setor'].apply(em_processo) if 'Setor' in pa.columns else pd.Series([False]*len(pa))
pa_disp=pa[~pa_proc_mask].copy()
for _,r in pa[pa_proc_mask].iterrows(): processo_rows.append(['Produto Acabado']+rec_padrao(r))
ws=wb.create_sheet('📦 Produto Acabado'); title_row(ws,1,len(HDRS_PA),f'PRODUTO ACABADO  |  {len(pa_disp)} registros  |  {DATA_STR}'); header_row(ws,2,HDRS_PA)
row=3
for linha,grupo in pa_disp.groupby('Linha',sort=True):
    n_reg=len(grupo); qtd_total=grupo['Estoque (UN)'].sum()
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=len(HDRS_PA))
    txt=f'  ▶  {linha or "(sem linha)"}   —   {n_reg} registros   |   Quantidade total: {qtd_total:,.0f}'.replace(',','.')
    c=ws.cell(row=row,column=1,value=txt); c.font=Font(name='Calibri',size=11,bold=True,color=GOLD); c.fill=PatternFill('solid',fgColor=LINHA_BG); c.alignment=Alignment(horizontal='left',vertical='center'); ws.row_dimensions[row].height=22
    row+=1
    for _,r in grupo.iterrows(): write_row(ws,row,HDRS_PA,['']+rec_padrao(r)); row+=1
for i,w in enumerate(WID_PA,1): ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes='A3'
familias=[('MP','🧪 Matéria Prima','MATÉRIA PRIMA','Matéria Prima'),('EMB','📦 Embalagem','EMBALAGEM','Embalagem'),('GRA','🛢️ Granel','GRANEL','Granel'),('APO','🔧 Apoio','APOIO','Apoio')]
for short,sheet_name,fam_upper,fam_nome in familias:
    sub=pd.read_csv(os.path.join(TMP,f'fam_{short}.csv'))
    proc_mask=sub['Setor'].apply(em_processo) if 'Setor' in sub.columns else pd.Series([False]*len(sub))
    disp=sub[~proc_mask].copy()
    for _,r in sub[proc_mask].iterrows(): processo_rows.append([fam_nome]+rec_padrao(r))
    wsf=wb.create_sheet(sheet_name); title_row(wsf,1,len(HDRS_FAM),f'{fam_upper}  |  {len(disp)} registros  |  {DATA_STR}'); header_row(wsf,2,HDRS_FAM)
    row=3
    for _,r in disp.iterrows(): write_row(wsf,row,HDRS_FAM,rec_padrao(r)); row+=1
    for i,w in enumerate(WID_FAM,1): wsf.column_dimensions[get_column_letter(i)].width=w
    wsf.freeze_panes='A3'
wsp=wb.create_sheet('⏳ Em Processo'); title_row(wsp,1,len(HDRS_PROC),f'EM PROCESSO (DOCA / PACKING)  |  {len(processo_rows)} registros  |  {DATA_STR}'); header_row(wsp,2,HDRS_PROC)
row=3; processo_rows.sort(key=lambda v:(str(v[0]),str(v[1])))
for valores in processo_rows:
    write_row(wsp,row,HDRS_PROC,valores)
    cf=wsp.cell(row=row,column=1); cf.font=Font(name='Calibri',size=9,bold=True,color=WHITE); cf.fill=PatternFill('solid',fgColor=PROC_BG); cf.alignment=Alignment(horizontal='center',vertical='center'); row+=1
for i,w in enumerate(WID_PROC,1): wsp.column_dimensions[get_column_letter(i)].width=w
wsp.freeze_panes='A3'
wb.save(OUT_XLSX)
print(f'[OK] Excel: {OUT_XLSX}'); print(f'  Abas: {wb.sheetnames}')
